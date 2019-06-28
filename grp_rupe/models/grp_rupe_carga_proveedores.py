# -*- coding: utf-8 -*-
##############################################################################
#
#    OpenERP, Open Source Enterprise Management Solution
#    GRP Estado Uruguay
#    Copyright (C) 2017 Quanam (ATEL SA., Uruguay)
#
#    This program is free software: you can redistribute it and/or modify
#    it under the terms of the GNU Affero General Public License as
#    published by the Free Software Foundation, either version 3 of the
#    License, or (at your option) any later version.
#
#    This program is distributed in the hope that it will be useful,
#    but WITHOUT ANY WARRANTY; without even the implied warranty of
#    MERCHANTABILITY or FITNESS FOR A PARTICULAR PURPOSE.  See the
#    GNU Affero General Public License for more details.
#
#    You should have received a copy of the GNU Affero General Public License
#    along with this program.  If not, see <http://www.gnu.org/licenses/>.
#
##############################################################################

import logging
import openerp
from openerp import api, exceptions, fields, models
from cStringIO import StringIO
import csv
import base64
import codecs, os
_logger = logging.getLogger(__name__)
import datetime
from openerp.exceptions import ValidationError


try:
    import openpyxl
    from openpyxl.cell import Cell
    from openpyxl import load_workbook
    from openpyxl.worksheet import Worksheet
    from openpyxl.drawing.image import Image
except:
    openpyxl = False


class grp_rupe_carga_proveedores(models.Model):
    _name = 'carga.proveedores'
    _rec_name = 'filename'

    filename = fields.Char("Archivo")
    archivo = fields.Binary(string="Seleccione archivo", required=True)
    errores_ids = fields.One2many('carga.proveedores.error', 'carga_id', 'Proveedores no creados')

    @api.multi
    def cargar_proveedores(self):
        vals={}
        try:
            if not openpyxl:
                raise ValidationError(u'La libreria de lectura de csv openpyxl no esta instalada')


            self.ensure_one()
            xlsx_io = StringIO()
            xlsx_io.write(base64.decodestring(self.archivo))
            wb = openpyxl.load_workbook(xlsx_io, read_only=True)
            if not wb._sheets:
                raise ValidationError(u'El archivo no contiene hojas')

            ws = wb._sheets[0]
            LINE = 0
            for row in ws.iter_rows():

                LINE += 1
                pool_staging_rupe = self.env['rupe.proveedores']
                prv_rupe_ids = pool_staging_rupe.search([('prv_cod_fiscal','=',row[0].value)])

                if not prv_rupe_ids:
                    error = {'numero': row[0].value,
                            'nombre':row[1].value,'carga_id': self.id , 'error':'No se encuentra en los proveedores RUPE'}
                    self.write({'errores_ids': [(0,0 , error)]})

                else:
                    partner = self.env['res.partner']

                    for p in prv_rupe_ids:
                        partner_ids = partner.search([('nro_doc_rupe','=',p.prv_cod_fiscal)])
                        if len(partner_ids) != 0 or p.incluido:
                            error = {'numero': row[0].value,'nombre':row[1].value,'carga_id': self.id , 'error':'Ya esta incluido en los proveedores'}
                            self.write({'errores_ids': [(0,0 , error)]})
                        else:
                            # Valores por defecto a traves del 'context'
                            vals['desde_import'] = True

                            vals['name'] = p.prv_denominacion_social
                            vals['street'] = p.prv_domicilio_fiscal
                            vals['city'] = p.prv_loc_fiscal_nombre
                            vals['email'] = p.prv_correo_electronico
                            vals['id_rupe'] = p.prv_id
                            vals['fecha_creacion_rupe'] = p.prv_crea_fecha
                            vals['version_rupe'] = p.prv_version
                            vals['estado_rupe'] = p.codigo_estado

                            tipo_doc = 'R'
                            if p.codigo_tipo_documento.strip() == 'RUT':
                                tipo_doc = 'R'
                            if p.codigo_tipo_documento.strip() == 'CI':
                                tipo_doc = 'CI'
                            if p.codigo_tipo_documento.strip() == 'PS':
                                tipo_doc = 'PS'
                            if p.codigo_tipo_documento.strip() == 'NIE':
                                tipo_doc = 'NIE'
                            if p.codigo_tipo_documento.strip() == 'CFE':
                                tipo_doc = 'CFE'

                            vals['tipo_doc_rupe'] = tipo_doc

                            vals['nro_doc_rupe'] = p.prv_cod_fiscal
                            vals['is_company'] = True
                            vals['fecha_modif_rupe'] = datetime.datetime.now().strftime("%Y-%m-%d")

                            vals['customer'] = False
                            vals['supplier'] = True

                            vals['id_staging'] = p.id

                            partner.create( vals)
                            p.write({'incluido':True})

            return True
        except Exception, e:
            raise ValidationError(u'No se pudo cargar planilla')





grp_rupe_carga_proveedores()

class grp_carga_proveedores_error(models.Model):
    _name = 'carga.proveedores.error'

    numero = fields.Char('Numero proveedor')
    nombre = fields.Char('Nombre')
    carga_id = fields.Many2one('carga.proveedores','Carga proveedores')
    error = fields.Char('Error')

grp_carga_proveedores_error()